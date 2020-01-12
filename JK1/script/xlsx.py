"""
============================
Author:virgir
Creation_time:2019-12-04
============================
"""
# coding=utf-8
import os
import openpyxl
from script import PATH


class Obj:
    pass


class Xlsx:
    def __init__(self, file_name=None, sheet=None):
        if file_name is None:
            self.file = os.path.join(PATH.xlsx_path, 'test.xlsx')
        else:
            self.file = os.path.join(PATH.xlsx_path, file_name)
        self.sheet = sheet

    def open_xlsx(self):
        self.ws = openpyxl.load_workbook(self.file)
        self.wb = self.ws[self.sheet]

    def read_xlsx(self):
        self.open_xlsx()
        data = []
        row = list(self.wb)
        title = [i.value for i in row[0]]
        for i in row[1:]:
            obj = Obj()
            li = [r.value for r in i]
            for o in list(zip(title, li)):
                setattr(obj, o[0], o[1])
            data.append(obj)
        self.ws.close()
        return data

    def write_xlsx(self, row, clunm, value):
        self.open_xlsx()
        self.wb.cell(row + 1, clunm, value)
        self.ws.save(self.file)
        self.ws.close()


if __name__ == '__main__':
    data = Xlsx(sheet='register')
    case = data.read_xlsx()
    print(case[0].case_id, case[0].data)
    print(case[1].case_id, case[1].data)
